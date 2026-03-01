#!/usr/bin/env python3
"""
Dataset Statistics Tool
-----------------------
Counts total images and samples (object instances) per class across
multiple dataset paths.

Usage:
    python dataset_stats.py <paths.json>
    python dataset_stats.py <paths.json> --xlsx report.xlsx

The JSON file must contain a top-level "paths" array. Each entry can be:
  - A plain string path
  - A dict with "path" (and optional "percentage")

For every directory the script finds all .json annotation files,
parses each one (expected to be a JSON array of objects with a "label"
field), and aggregates the counts.
"""

import argparse
import json
import os
import sys
from collections import defaultdict
from pathlib import Path

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff", ".webp"}


def resolve_path(entry):
    """Return the directory path from a string or dict entry."""
    if isinstance(entry, str):
        return entry
    if isinstance(entry, dict) and "path" in entry:
        return entry["path"]
    return None


def count_images(directory: str) -> int:
    """Count image files in the directory (non-recursive)."""
    count = 0
    try:
        for f in os.listdir(directory):
            if os.path.splitext(f)[1].lower() in IMAGE_EXTENSIONS:
                count += 1
    except FileNotFoundError:
        pass
    return count


def count_samples(directory: str):
    """
    Parse all .json files in the directory and count object instances
    grouped by label (class).

    Returns:
        dict[str, int]: label -> count
    """
    class_counts = defaultdict(int)
    try:
        files = sorted(os.listdir(directory))
    except FileNotFoundError:
        return class_counts

    for fname in files:
        if not fname.lower().endswith(".json"):
            continue
        fpath = os.path.join(directory, fname)
        try:
            with open(fpath, "r") as fh:
                data = json.load(fh)
        except (json.JSONDecodeError, IOError):
            continue

        if not isinstance(data, list):
            continue

        for obj in data:
            label = obj.get("label", "unknown")
            if label == "":
                label = "unknown"
            class_counts[label] += 1

    return class_counts


def collect_data(paths):
    """Scan all paths and return structured results + grand totals."""
    results = []
    grand_images = 0
    grand_samples = 0
    grand_classes = defaultdict(int)
    all_labels = set()

    for entry in paths:
        dirpath = resolve_path(entry)
        if dirpath is None:
            continue

        exists = os.path.isdir(dirpath)
        images = count_images(dirpath) if exists else 0
        classes = dict(count_samples(dirpath)) if exists else {}
        total_samples = sum(classes.values())

        all_labels.update(classes.keys())
        grand_images += images
        grand_samples += total_samples
        for label, cnt in classes.items():
            grand_classes[label] += cnt

        results.append({
            "path": dirpath,
            "exists": exists,
            "images": images,
            "samples": total_samples,
            "classes": classes,
        })

    return results, dict(grand_classes), grand_images, grand_samples, sorted(all_labels)


def print_terminal(results, grand_classes, grand_images, grand_samples, all_labels):
    """Pretty-print results to the terminal."""
    separator = "─" * 90

    print(f"\n{'DATASET STATISTICS':^90}")
    print(f"{'=' * 90}\n")

    for r in results:
        print(f"  📂 {r['path']}")
        if not r["exists"]:
            print(f"     ⚠  Directory not found!")
        else:
            print(f"     Images : {r['images']}")
            print(f"     Samples: {r['samples']}")
            if r["classes"]:
                max_label_len = max(len(l) for l in r["classes"])
                for label in sorted(r["classes"]):
                    print(f"       • {label:<{max_label_len}} : {r['classes'][label]}")
        print(f"  {separator}")

    print(f"\n{'GRAND TOTALS':^90}")
    print(f"{'=' * 90}")
    print(f"  Total images  : {grand_images}")
    print(f"  Total samples : {grand_samples}")
    if grand_classes:
        print(f"\n  Samples by class:")
        max_label_len = max(len(l) for l in grand_classes)
        for label in sorted(grand_classes):
            pct = grand_classes[label] / grand_samples * 100 if grand_samples else 0
            print(f"    • {label:<{max_label_len}} : {grand_classes[label]:>8}  ({pct:5.1f}%)")
    print(f"{'=' * 90}\n")


def export_xlsx(filepath, results, grand_classes, grand_images, grand_samples, all_labels):
    """Export results to a nicely formatted .xlsx file."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Dataset Statistics"

    # ── Styles ────────────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    subheader_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    total_font = Font(name="Calibri", bold=True, size=11)
    total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    grand_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    grand_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    normal_font = Font(name="Calibri", size=11)
    warning_font = Font(name="Calibri", size=11, color="CC0000", italic=True)
    pct_font = Font(name="Calibri", size=10, color="666666")
    thin_border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    # Alternating row colors
    even_fill = PatternFill(start_color="EDF2F9", end_color="EDF2F9", fill_type="solid")
    odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # ── Build columns: Path | Images | Samples | label1 | label2 | ... ──
    fixed_cols = ["Path", "Images", "Samples"]
    headers = fixed_cols + [lbl.capitalize() for lbl in all_labels]

    # ── Title row ─────────────────────────────────────────────────
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
    title_cell = ws.cell(row=row, column=1, value="Dataset Statistics Report")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30
    row += 1

    # ── Header row ────────────────────────────────────────────────
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    ws.row_dimensions[row].height = 22
    row += 1

    # ── Data rows ─────────────────────────────────────────────────
    for data_idx, r in enumerate(results):
        row_fill = even_fill if data_idx % 2 == 0 else odd_fill

        # Path
        cell = ws.cell(row=row, column=1, value=r["path"])
        cell.font = warning_font if not r["exists"] else normal_font
        cell.alignment = left
        cell.fill = row_fill
        cell.border = thin_border

        if not r["exists"]:
            cell = ws.cell(row=row, column=2, value="NOT FOUND")
            cell.font = warning_font
            cell.alignment = center
            cell.fill = row_fill
            cell.border = thin_border
            for ci in range(3, len(headers) + 1):
                c = ws.cell(row=row, column=ci, value="")
                c.fill = row_fill
                c.border = thin_border
        else:
            # Images
            cell = ws.cell(row=row, column=2, value=r["images"])
            cell.font = normal_font
            cell.alignment = center
            cell.fill = row_fill
            cell.border = thin_border
            cell.number_format = "#,##0"

            # Samples
            cell = ws.cell(row=row, column=3, value=r["samples"])
            cell.font = normal_font
            cell.alignment = center
            cell.fill = row_fill
            cell.border = thin_border
            cell.number_format = "#,##0"

            # Per-class counts
            for lbl_idx, lbl in enumerate(all_labels):
                val = r["classes"].get(lbl, 0)
                cell = ws.cell(row=row, column=4 + lbl_idx, value=val if val > 0 else "")
                cell.font = normal_font
                cell.alignment = center
                cell.fill = row_fill
                cell.border = thin_border
                if val > 0:
                    cell.number_format = "#,##0"

        row += 1

    # ── Grand Totals row ──────────────────────────────────────────
    row += 1  # blank row
    cell = ws.cell(row=row, column=1, value="GRAND TOTALS")
    cell.font = grand_font
    cell.fill = grand_fill
    cell.alignment = left
    cell.border = thin_border

    cell = ws.cell(row=row, column=2, value=grand_images)
    cell.font = grand_font
    cell.fill = grand_fill
    cell.alignment = center
    cell.border = thin_border
    cell.number_format = "#,##0"

    cell = ws.cell(row=row, column=3, value=grand_samples)
    cell.font = grand_font
    cell.fill = grand_fill
    cell.alignment = center
    cell.border = thin_border
    cell.number_format = "#,##0"

    for lbl_idx, lbl in enumerate(all_labels):
        val = grand_classes.get(lbl, 0)
        cell = ws.cell(row=row, column=4 + lbl_idx, value=val)
        cell.font = grand_font
        cell.fill = grand_fill
        cell.alignment = center
        cell.border = thin_border
        cell.number_format = "#,##0"

    # ── Percentage row ────────────────────────────────────────────
    row += 1
    cell = ws.cell(row=row, column=1, value="% of total")
    cell.font = Font(name="Calibri", bold=True, size=10, color="666666")
    cell.fill = total_fill
    cell.alignment = left
    cell.border = thin_border

    for ci in [2, 3]:
        c = ws.cell(row=row, column=ci, value="")
        c.fill = total_fill
        c.border = thin_border

    for lbl_idx, lbl in enumerate(all_labels):
        val = grand_classes.get(lbl, 0)
        pct = val / grand_samples if grand_samples else 0
        cell = ws.cell(row=row, column=4 + lbl_idx, value=pct)
        cell.font = pct_font
        cell.fill = total_fill
        cell.alignment = center
        cell.border = thin_border
        cell.number_format = "0.0%"

    # ── Column widths ─────────────────────────────────────────────
    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    for i in range(len(all_labels)):
        col_letter = get_column_letter(4 + i)
        ws.column_dimensions[col_letter].width = 14

    # ── Freeze panes (header stays visible while scrolling) ──────
    ws.freeze_panes = "A3"

    # ── Auto-filter ───────────────────────────────────────────────
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{row}"

    wb.save(filepath)
    return filepath


def main():
    parser = argparse.ArgumentParser(
        description="Count images and samples (by class) in dataset paths."
    )
    parser.add_argument(
        "config",
        help="JSON file containing a 'paths' array.",
    )
    parser.add_argument(
        "--xlsx",
        nargs="?",
        const="auto",
        default=None,
        metavar="FILE",
        help="Export results to an .xlsx file. If no filename is given, "
             "auto-generates one from the config name.",
    )
    args = parser.parse_args()

    with open(args.config, "r") as fh:
        config = json.load(fh)

    paths = config.get("paths", [])
    if not paths:
        print("No paths found in the config file.", file=sys.stderr)
        sys.exit(1)

    # ── Collect ───────────────────────────────────────────────────
    results, grand_classes, grand_images, grand_samples, all_labels = collect_data(paths)

    # ── Terminal output ───────────────────────────────────────────
    print_terminal(results, grand_classes, grand_images, grand_samples, all_labels)

    # ── XLSX export ───────────────────────────────────────────────
    if args.xlsx is not None:
        if args.xlsx == "auto":
            base = Path(args.config).stem
            xlsx_path = f"{base}_report.xlsx"
        else:
            xlsx_path = args.xlsx

        # Warn about wrong extension
        if not xlsx_path.lower().endswith(".xlsx"):
            print(f"  ⚠  Warning: '{xlsx_path}' does not end with .xlsx — "
                  f"Excel may not open it correctly.", file=sys.stderr)
            xlsx_path = os.path.splitext(xlsx_path)[0] + ".xlsx"
            print(f"     Saving as: {xlsx_path}", file=sys.stderr)

        out = export_xlsx(xlsx_path, results, grand_classes, grand_images, grand_samples, all_labels)
        print(f"  📊 Exported to: {out}\n")


if __name__ == "__main__":
    main()
